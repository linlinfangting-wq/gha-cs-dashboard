#!/usr/bin/env python3
"""
GHA 客服数据报告生成器
生成 Excel 文件，包含数据明细 + 4张可视化图表
用法: python3 generate_report.py
输出: report_YYYY-MM-DD.xlsx
"""

import os
import csv
import sys
from datetime import datetime
from collections import Counter

# 检查依赖
try:
    import pandas as pd
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ 缺少依赖: {e}")
    print("请运行: pip3 install pandas matplotlib openpyxl")
    sys.exit(1)

CSV_PATH = os.path.join(os.path.dirname(__file__), 'records.csv')
OUT_DIR = os.path.dirname(__file__)

# ── 字体配置（中文支持）──
def get_chinese_font():
    candidates = [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/Library/Fonts/Arial Unicode MS.ttf',
    ]
    for path in candidates:
        if os.path.exists(path):
            return fm.FontProperties(fname=path)
    return fm.FontProperties()

FONT = get_chinese_font()
plt.rcParams['axes.unicode_minus'] = False

# ── 品牌色 ──
COLOR_GOLD   = '#DA9F59'
COLOR_NAVY   = '#141432'
COLOR_BLUE   = '#8BBCD9'
COLOR_CORAL  = '#F69B6F'
COLOR_GREEN  = '#5D8C6A'
COLOR_PURPLE = '#300B5C'
COLOR_IVORY  = '#FAFAF8'

PALETTE = [COLOR_GOLD, COLOR_BLUE, COLOR_CORAL, COLOR_GREEN, COLOR_PURPLE,
           '#B7CBD5', '#C4A882', '#7A9898', '#E8C880', '#9090B8']


def load_data():
    if not os.path.exists(CSV_PATH):
        print("❌ records.csv 不存在"); sys.exit(1)
    df = pd.read_csv(CSV_PATH, encoding='utf-8')
    df = df[df['id'].notna() & (df['id'].astype(str).str.strip() != '')]
    if df.empty:
        print("❌ 暂无数据，请先录入记录"); sys.exit(1)
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    return df


def save_chart(fig, name):
    path = os.path.join(OUT_DIR, f'_tmp_{name}.png')
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return path


# ── 图1：问题分类分布（横向条形图）──
def chart_category(df):
    counts = df['category'].value_counts()
    fig, ax = plt.subplots(figsize=(8, max(3, len(counts)*0.5 + 1)))
    bars = ax.barh(counts.index[::-1], counts.values[::-1],
                   color=PALETTE[:len(counts)], height=0.6)
    for bar, val in zip(bars, counts.values[::-1]):
        ax.text(val + 0.1, bar.get_y() + bar.get_height()/2,
                str(val), va='center', fontproperties=FONT, fontsize=10)
    ax.set_xlabel('咨询量', fontproperties=FONT)
    ax.set_title('问题分类分布', fontproperties=FONT, fontsize=14, pad=12)
    ax.tick_params(axis='y')
    for label in ax.get_yticklabels():
        label.set_fontproperties(FONT)
    for label in ax.get_xticklabels():
        label.set_fontproperties(FONT)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    return save_chart(fig, 'category')


# ── 图2：渠道来源分布（饼图）──
def chart_channel(df):
    counts = df['channel'].value_counts()
    fig, ax = plt.subplots(figsize=(6, 6))
    wedges, texts, autotexts = ax.pie(
        counts.values, labels=counts.index,
        colors=PALETTE[:len(counts)],
        autopct='%1.1f%%', startangle=90,
        pctdistance=0.75,
        wedgeprops=dict(width=0.55)  # donut
    )
    for t in texts:
        t.set_fontproperties(FONT)
        t.set_fontsize(11)
    for a in autotexts:
        a.set_fontsize(10)
        a.set_fontweight('bold')
    ax.set_title('渠道来源分布', fontproperties=FONT, fontsize=14, pad=12)
    fig.tight_layout()
    return save_chart(fig, 'channel')


# ── 图3：处理状态分布（条形图）──
def chart_status(df):
    status_order = ['已解决', '处理中', '已升级', '待处理']
    counts = df['status'].value_counts().reindex(status_order).dropna()
    status_colors = {
        '已解决': COLOR_GREEN, '处理中': COLOR_GOLD,
        '已升级': COLOR_CORAL, '待处理': COLOR_BLUE
    }
    colors = [status_colors.get(s, COLOR_NAVY) for s in counts.index]
    fig, ax = plt.subplots(figsize=(6, 4))
    bars = ax.bar(counts.index, counts.values, color=colors, width=0.5)
    for bar, val in zip(bars, counts.values):
        ax.text(bar.get_x() + bar.get_width()/2, val + 0.1,
                str(val), ha='center', va='bottom', fontproperties=FONT, fontsize=11)
    ax.set_ylabel('数量', fontproperties=FONT)
    ax.set_title('处理状态分布', fontproperties=FONT, fontsize=14, pad=12)
    for label in ax.get_xticklabels():
        label.set_fontproperties(FONT)
    for label in ax.get_yticklabels():
        label.set_fontproperties(FONT)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    return save_chart(fig, 'status')


# ── 图4：每日咨询量趋势（折线图）──
def chart_trend(df):
    daily = df.groupby(df['date'].dt.date).size().reset_index(name='count')
    fig, ax = plt.subplots(figsize=(9, 4))
    ax.plot(range(len(daily)), daily['count'], color=COLOR_GOLD,
            linewidth=2, marker='o', markersize=6)
    ax.fill_between(range(len(daily)), daily['count'],
                    color=COLOR_GOLD, alpha=0.1)
    ax.set_xticks(range(len(daily)))
    ax.set_xticklabels([str(d) for d in daily['date']],
                       rotation=30, ha='right')
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontproperties(FONT)
    ax.set_ylabel('咨询量', fontproperties=FONT)
    ax.set_title('每日咨询量趋势', fontproperties=FONT, fontsize=14, pad=12)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    return save_chart(fig, 'trend')


# ── 图5：高频关键词 Top10 ──
def chart_keywords(df):
    all_kw = []
    for row in df['keywords'].dropna():
        all_kw.extend([k.strip() for k in str(row).split(',') if k.strip()])
    if not all_kw:
        return None
    top = Counter(all_kw).most_common(10)
    labels = [t[0] for t in top][::-1]
    values = [t[1] for t in top][::-1]
    fig, ax = plt.subplots(figsize=(8, max(3, len(labels)*0.45 + 1)))
    bars = ax.barh(labels, values, color=COLOR_BLUE, height=0.6)
    for bar, val in zip(bars, values):
        ax.text(val + 0.05, bar.get_y() + bar.get_height()/2,
                str(val), va='center', fontproperties=FONT, fontsize=10)
    ax.set_xlabel('出现次数', fontproperties=FONT)
    ax.set_title('高频关键词 Top10', fontproperties=FONT, fontsize=14, pad=12)
    for label in ax.get_yticklabels() + ax.get_xticklabels():
        label.set_fontproperties(FONT)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    return save_chart(fig, 'keywords')


# ── Excel 写入 ──
def write_excel(df, chart_paths, out_path):
    wb = Workbook()

    # ── Sheet 1: 数据明细 ──
    ws1 = wb.active
    ws1.title = '数据明细'
    header_fill = PatternFill(fill_type='solid', fgColor='141432')
    header_font = Font(color='FFFFFF', bold=True, name='Microsoft YaHei', size=10)
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_names = ['ID', '日期', '时间', '渠道', '问题类别', '子类别',
                 '关键词', '描述摘要', '处理状态', '备注']
    col_widths = [6, 12, 8, 10, 14, 14, 28, 40, 10, 30]

    for c, (name, width) in enumerate(zip(col_names, col_widths), 1):
        cell = ws1.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws1.column_dimensions[get_column_letter(c)].width = width

    status_colors = {
        '已解决': 'D4EDDA', '处理中': 'FFF3CD',
        '已升级': 'F8D7DA', '待处理': 'D1ECF1'
    }
    for r, row in enumerate(df.itertuples(), 2):
        vals = [row.id, str(row.date)[:10] if pd.notna(row.date) else '',
                getattr(row, 'time', ''), row.channel, row.category,
                row.subcategory, row.keywords, row.description,
                row.status, row.notes]
        for c, val in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=str(val) if pd.notna(val) else '')
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=(c in [7, 8, 10]))
            # 状态列着色
            if c == 9:
                sc = status_colors.get(str(val), 'FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor=sc)
        ws1.row_dimensions[r].height = 22
    ws1.row_dimensions[1].height = 24
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: 数据图表 ──
    ws2 = wb.create_sheet('数据图表')
    ws2.sheet_view.showGridLines = False

    # 标题
    ws2['B2'] = 'GHA 会员客服数据分析报告'
    ws2['B2'].font = Font(name='Microsoft YaHei', bold=True, size=16, color='141432')
    ws2['B3'] = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}  |  共 {len(df)} 条记录'
    ws2['B3'].font = Font(name='Microsoft YaHei', size=10, color='888888')

    # 嵌入图表图片
    positions = ['B5', 'L5', 'B32', 'L32', 'B59']
    for path, pos in zip(chart_paths, positions):
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width = int(img.width * 0.72)
            img.height = int(img.height * 0.72)
            ws2.add_image(img, pos)

    # ── Sheet 3: 统计汇总 ──
    ws3 = wb.create_sheet('统计汇总')
    ws3['B2'] = '统计项目'
    ws3['C2'] = '数值'
    ws3['B2'].font = Font(bold=True, color='FFFFFF')
    ws3['C2'].font = Font(bold=True, color='FFFFFF')
    ws3['B2'].fill = PatternFill(fill_type='solid', fgColor='141432')
    ws3['C2'].fill = PatternFill(fill_type='solid', fgColor='141432')

    stats = [
        ('总咨询量', len(df)),
        ('已解决', len(df[df['status'] == '已解决'])),
        ('解决率', f"{len(df[df['status']=='已解决'])/len(df)*100:.1f}%"),
        ('处理中', len(df[df['status'] == '处理中'])),
        ('已升级', len(df[df['status'] == '已升级'])),
        ('涉及渠道数', df['channel'].nunique()),
        ('问题类别数', df['category'].nunique()),
        ('统计区间', f"{df['date'].min().strftime('%Y-%m-%d') if pd.notna(df['date'].min()) else '-'} ~ {df['date'].max().strftime('%Y-%m-%d') if pd.notna(df['date'].max()) else '-'}"),
    ]
    for i, (label, value) in enumerate(stats, 3):
        ws3[f'B{i}'] = label
        ws3[f'C{i}'] = value
        ws3[f'B{i}'].font = Font(name='Microsoft YaHei', size=10)
        ws3[f'C{i}'].font = Font(name='Microsoft YaHei', size=10, bold=True, color='DA9F59')
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 22

    wb.save(out_path)
    print(f'✅ 报告已生成：{out_path}')


def cleanup_tmp():
    for name in ['category', 'channel', 'status', 'trend', 'keywords']:
        p = os.path.join(OUT_DIR, f'_tmp_{name}.png')
        if os.path.exists(p):
            os.remove(p)


if __name__ == '__main__':
    df = load_data()
    print(f'📊 共 {len(df)} 条记录，开始生成图表...')

    charts = [
        chart_category(df),
        chart_channel(df),
        chart_status(df),
        chart_trend(df),
        chart_keywords(df),
    ]

    out_path = os.path.join(OUT_DIR, f'GHA客服报告_{datetime.now().strftime("%Y-%m-%d")}.xlsx')
    write_excel(df, charts, out_path)
    cleanup_tmp()
    print(f'📁 文件位置：{out_path}')
